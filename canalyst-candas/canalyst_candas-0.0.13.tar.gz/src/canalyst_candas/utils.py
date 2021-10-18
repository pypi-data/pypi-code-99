"""
Utilities module
"""
from typing import Dict
import pandas as pd
from python_graphql_client import GraphqlClient
import numpy as np
from urllib.parse import quote_plus
from io import StringIO
from canalyst_candas import settings
import re
import boto3
from boto3.session import Session
import os
import string
from os import path
import requests
import time
import json
from openpyxl import load_workbook, styles
import datetime
import csv
from canalyst_candas.version import __version__ as version

from canalyst_candas.configuration.config import Config, resolve_config
import urllib3

urllib3.disable_warnings()


CSIN_URL = "api/equity-model-series/?company_ticker_bloomberg={ticker}"
PERIODS_URL = (
    "api/equity-model-series/{csin}/equity-models/{version}/historical-periods/"
)
SCENARIO_URL = "api/equity-model-series/{csin}/equity-models/{version}/scenarios/"
SCENARIO_URL_FORECAST = (
    "api/equity-model-series/{csin}/equity-models/"
    "{version}/scenarios/{scenario_id}/forecast-periods/"
)

HISTORICAL_PERIODS_TO_FETCH = 24


def set_api_key(key, key_file_path):
    with open(key_file_path) as f:
        keys_json = json.load(f)

    keys_json["canalyst_api_key"] = key

    with open(key_file_path, "w") as f:
        json.dump(keys_json, f)


def get_api_headers(canalyst_api_key: str) -> Dict[str, str]:
    """
    Return the authorization bearer header to use for API requests and user agent
    """
    return {
        "Authorization": f"Bearer {canalyst_api_key}",
        "User-Agent": f"canalyst-sdk-{version}",
    }


class LogFile:

    # to be refactored to log try: except: errors

    # the idea of this class is to help with user debug ...
    def __init__(self, default_dir: str = settings.DEFAULT_DIR, verbose: bool = False):
        self.default_dir = default_dir
        self.verbose = verbose
        tm = datetime.datetime.now()
        self.log_file_name = f"{default_dir}/candas_logfile.csv"

        if not os.path.isfile(self.log_file_name):
            rows = [["timestamp", "action"], [tm, "initiate logfile"]]
            with open(self.log_file_name, "w", newline="") as csvfile:
                csvwriter = csv.writer(csvfile)
                csvwriter.writerows(rows)

    def write(self, text):
        if self.verbose is True:
            print(text)
        tm = datetime.datetime.now()
        rows = [tm, text]
        with open(self.log_file_name, "a", newline="") as csvfile:
            csvwriter = csv.writer(csvfile)
            csvwriter.writerow(rows)

    def read(self):
        df = pd.read_csv(self.log_file_name)
        return df


# helper class to get data from S3
class Getter:
    def __init__(self, config: Config):
        self.config = config
        self.log = LogFile(default_dir=self.config.default_dir)
        self.s3_client = boto3.client(
            "s3",
            aws_access_key_id=self.config.s3_access_key_id,
            aws_secret_access_key=self.config.s3_secret_key,
        )

    def get_s3_file(self, to_filename, from_filename):
        try:
            with open(to_filename, "wb+") as f:
                self.s3_client.download_fileobj("candas", from_filename, f)
            print(f"Downloaded to {to_filename}. ")
        except:
            self.log.write("Get S3 file failed")
        return

    def get_file_from_s3(self, file_name):
        try:
            csv_obj = self.s3_client.get_object(Bucket="candas", Key=file_name)
            body = csv_obj["Body"].read()
        except:
            self.log.write("Missing file from s3." + file_name)
            return None
        return body

    def get_zip_csv_from_s3(self, file_name):
        d = self.s3_client.get_object(Bucket="candas", Key=file_name)
        import io

        buffer = io.BytesIO(d["Body"].read())
        import zipfile

        z = zipfile.ZipFile(buffer)
        body = z.open(z.namelist()[0]).read().decode("utf-8")
        df = pd.read_csv(io.StringIO(body), low_memory=False)
        return df

    def get_csv_from_s3(self, file_name):
        try:
            csv_obj = self.s3_client.get_object(Bucket="candas", Key=file_name)
            body = csv_obj["Body"]
            csv_string = body.read().decode("utf-8")
            df = pd.read_csv(StringIO(csv_string))
        except:
            self.log.write("Missing csv from s3." + file_name)
            return None
        return df

    def get_file(self, ticker, type):

        file_ticker = ticker.split(" ")[0]
        path_name = f"DATA/{file_ticker}/{ticker}_{type}.csv"
        df = self.get_csv_from_s3(path_name)

        if df is None:
            if path.exists(f"{path_name}"):
                df = pd.read_csv(f"{path_name}")

        if df is None:
            return None

        keep_cols = []
        for col in df.columns:
            if "Unnamed" not in col:
                keep_cols.append(col)
        df = df[keep_cols]

        return df


def save_guidance_csv(df, ticker, filename):
    pathname = filename  # CWD + "\\DATA\\" + filename + "\\"  # .replace(r,'')

    for i, row in df.iterrows():
        for col in row.index:
            val = row[col]
            val = str(val)

            df = df.iloc[3:, 1:]
            df = df.dropna(axis=1, how="all")

            df.columns = df.iloc[0]
            df = df.drop(df.index[0])
            df.to_csv(pathname + "/" + ticker + "_guidance.csv")
            print(ticker + " guidance done")
            return df


def filter_dataset(
    df,
    time_series_name="",
    period_name="",
    is_driver="",
    pivot=False,
    mrq=False,
    period_duration_type="",
    is_historical="",
    n_periods="",
    mrq_notation=False,
    unit_type="",
    category="",
    warning=True,
):

    if (
        period_duration_type == ""
        and period_name == ""
        and warning is True
        and mrq is False
    ):
        print(
            "Warning: Returning both FY and Q.  Use period_duration_type = 'fiscal_quarter' or "
            "period_duration_type = 'fiscal_year' to filter."
        )

    if mrq is True:
        mrq = df["MRQ"].iloc[0]
        df = df.loc[df["period_name"] == df["MRQ"]]

    df = df.sort_values(["period_end_date", "name_index"])

    if category != "":
        df = df.loc[df["category"] == category]

    if period_name != "":
        df = df.loc[df["period_name"] == period_name]

    if period_duration_type != "":
        df = df.loc[df["period_duration_type"] == period_duration_type]

    if is_driver != "":
        df = df.loc[df["is_driver"] == is_driver]

    if period_name != "":
        df = df.loc[df["period_name"] == period_name]

    if is_historical != "":
        df = df.loc[df["is_historical"] == is_historical]

    if unit_type != "":
        df = df.loc[df["unit_type"] == unit_type]

    if type(n_periods) is range:  # is this a range
        list_numbers = []
        list_numbers = list(n_periods)
        df = df.groupby(["ticker", "time_series_name"]).take(list_numbers).reset_index()
    elif type(n_periods) is int:
        if n_periods > 0:
            df = (
                df.groupby(["ticker", "time_series_name"]).tail(n_periods).reset_index()
            )
        else:
            df = (
                df.groupby(["ticker", "time_series_name"])
                .head(-1 * n_periods)
                .reset_index()
            )

    df["period_name_sorted"] = (
        df["period_name"].str.split("-").str[1]
        + df["period_name"].str.split("-").str[0]
    )

    if time_series_name == "":
        if pivot == True:
            return pivot_df(df, mrq_notation)
        return df

    if type(time_series_name) is not list and time_series_name != "":
        df1 = df[
            df["time_series_name"].str.contains(
                time_series_name, flags=re.IGNORECASE, regex=True
            )
        ]

        df2 = df[
            df["time_series_description"].str.contains(
                time_series_name, flags=re.IGNORECASE, regex=True
            )
        ]

        df = pd.concat([df1, df2])

        df = df.sort_values("period_end_date")

        if pivot is True:
            df = pivot_df(df, mrq_notation)
        return df

    if type(time_series_name) is list:
        list_df = []
        for item in time_series_name:
            df_copy = df.loc[df["time_series_name"] == item]
            df_copy = df_copy.sort_values("period_end_date")
            list_df.append(df_copy)
        df = pd.concat(list_df)

    if pivot is True:
        df = pivot_df(df, mrq_notation)

    return df


def send_scenario(
    ticker, data, auth_headers, csin="", latest_version="", mds_host=settings.MDS_HOST
):

    if csin == "" or latest_version == "":
        csin, latest_version = get_company_info_from_ticker(
            ticker, auth_headers, mds_host
        )
    scenario_url = (
        f"{mds_host}/{SCENARIO_URL.format(csin=csin, version=latest_version)}"
    )
    # print("scenario_url: " + scenario_url)

    scenario_response = post_request(url=scenario_url, headers=auth_headers, json=data)
    # print(scenario_response.text)

    return scenario_response


def make_scenario_data(feature_name, feature_period, feature_value, scenario_name):
    data = {
        "changes": [
            {
                "time_series": feature_name,
                "period": feature_period,
                "value_expression": {"type": "literal", "value": feature_value},
            }
        ],
        "name": scenario_name,
    }
    return data


def map_scenario_urls(json):
    scenario_map = {}
    results = json.get("results", {})
    for result in results:
        name = result.get("name")
        url = result.get("self")
        scenario_map[name] = url

    return scenario_map


def get_drivers_from_model(
    ticker,
    auth_headers,
    default_dir=settings.DEFAULT_DIR,
    mds_host=settings.MDS_HOST,
    wp_host=settings.WP_HOST,
):
    # check if the model exists if so use it.
    file_ticker = ticker.split(" ")[0]

    log = LogFile(default_dir=default_dir)

    get_excel_model(ticker, auth_headers, default_dir, mds_host, wp_host)

    path_name = f"{default_dir}/DATA/{file_ticker}/"

    cache_files = []
    _, files = scandir(path_name, [".xlsx"])
    for filename in files:
        if "data" not in filename:
            cache_files.append(filename)
    list_for = []
    for f_path in cache_files:
        if "~" in f_path or ".csv" in f_path:
            continue

        workbook = load_workbook(f_path, data_only=True)
        model_sheet = workbook["Model"]
        list_names = [
            defined_name for defined_name in workbook.defined_names.definedName
        ]

        dict_names = {}
        for named_range in list_names:
            if "Model!" in named_range.attr_text and ":" in named_range.attr_text:
                x = str(named_range.attr_text).replace("Model!", "")
                x = x.split(":")[0]
                x = x.replace("$", "")
                x = x.replace("A", "")
                x = x.replace("C", "")
                x = x.replace("B", "")

                if x != "":
                    dict_names[named_range.name] = x

        df_index = pd.DataFrame(dict_names, index=[0]).T
        df_index = df_index.reset_index()
        df_index.columns = ["time_series_name", "index"]
        df_index["index"] = df_index["index"].astype(int)
        df_index = df_index.sort_values("index")

        for named_range in list_names:
            if "LatestMRQ" in str(named_range):
                mrq = str(named_range.value)
                mrq = mrq.replace('"', "")
            if "Common_ColumnHeader" in str(named_range):
                col_hdr = str(named_range.value)
                col_hdr = str(named_range.value).split("!")[1]
                col_hdr = col_hdr.replace('"', "")

        col_hdr = col_hdr.split(":")[0]
        col_hdr = col_hdr.replace("$", "")
        col_hdr = "A" + col_hdr + ":DA" + col_hdr

        def n2a(n, b=string.ascii_uppercase):
            d, m = divmod(n, len(b))
            return n2a(d - 1, b) + b[m] if d else b[m]

        driver_sheet = workbook["Drivers"]

        cells = list(driver_sheet["A5":"DA5"])
        headers = [cell.value for cell in cells[0]]
        for index, item in enumerate(headers):
            if item == mrq:
                col_loc_i = index

        # not 'z_' in str(named_range) and
        named_ranges = {}
        for named_range in list_names:
            # and not 'z_' in str(named_range)
            if (
                "MO_" in str(named_range.name)
                and ":" in str(named_range.attr_text)
                and not "MO_Section" in str(named_range.name)
                and not "MO_SPT_" in str(named_range.name)
                and not "MO_SubSection_" in str(named_range.name)
                and not "MO_SNA_" in str(named_range.name)
                and not "MO_Checks_" in str(named_range.name)
                and not "MO_Common_" in str(named_range.name)
            ):
                z = named_range.attr_text.split(":")
                z[0] = z[0].split("!")[1]
                named_ranges[named_range.name] = z[0]

        list_df = []

        for defined_name in named_ranges:
            row = named_ranges[defined_name].replace("$", "")
            try:
                list_values = [
                    val
                    for val in model_sheet.iter_rows(min_row=int(row), max_row=int(row))
                ][0]
            except:
                log.write(f"error: {defined_name}")
                continue

            dict_values = {}

            for index, item in enumerate(list_values):
                i_toggle = 0
                if type(item.value) == float or type(item.value) == int:
                    dict_values = {}
                    num = (
                        "$"
                        + n2a(index)
                        + str(named_ranges[defined_name]).replace("$", "")
                    )
                    font_color = item.font.color
                    if font_color and font_color.type == "rgb":
                        color = font_color.rgb
                    elif font_color and font_color.type == "indexed":
                        color = styles.colors.COLOR_INDEX[font_color.indexed]
                    else:
                        color = styles.colors.COLOR_INDEX[0]  # black

                    if color.endswith("FF0000"):
                        if index <= col_loc_i:
                            dict_values["period_name"] = headers[index - 1]
                            dict_values["cell"] = num
                            dict_values["actual"] = 0
                            dict_values["historical"] = 1
                        else:
                            dict_values["period_name"] = headers[index - 1]
                            dict_values["cell"] = num
                            dict_values["actual"] = 0
                            dict_values["historical"] = 0
                        i_toggle = 1
                    else:
                        if index <= col_loc_i:
                            dict_values["period_name"] = headers[index - 1]
                            dict_values["cell"] = num
                            dict_values["actual"] = 1
                            dict_values["historical"] = 1
                        else:
                            dict_values["period_name"] = headers[index - 1]
                            dict_values["cell"] = num
                            dict_values["actual"] = 1
                            dict_values["historical"] = 0
                    df = pd.DataFrame(dict_values, index=[0])
                    df["named_range"] = defined_name

                    if i_toggle == 1:
                        list_df.append(df)

        df = pd.concat(list_df).sort_values(["named_range", "cell"])
        df["filename"] = f_path

        workbook.close()
        list_for.append(df)
    return pd.concat(list_for), df_index


def get_scenarios(ticker, auth_headers, mds_host=settings.MDS_HOST):
    if ticker == "":
        ticker = ticker
    url = f"{mds_host}/api/scenarios/"
    csin, latest_version = get_company_info_from_ticker(ticker, auth_headers, mds_host)
    response = requests.get(url, headers=auth_headers, verify=False)
    res = response.json().get("results")
    list_out = []
    for i in res:
        dict_out = {}
        dict_out["model_name"] = i["name"]
        dict_out["model_csin"] = i["equity_model"]["equity_model_series"]["csin"]
        df = pd.DataFrame.from_dict(dict_out, orient="index").T
        list_out.append(df)
    df_data = pd.concat(list_out)
    df_data = df_data.loc[df_data["model_csin"] == csin]
    return df_data


def post_request(url, headers, json):
    response = requests.post(url=url, headers=headers, json=json, verify=False)
    if check_throttle_limit(response):
        response = requests.post(url=url, headers=headers, json=json, verify=False)

    return response


def get_company_info_from_ticker(ticker, auth_headers, mds_host=settings.MDS_HOST):

    company_url = f"{mds_host}/{CSIN_URL.format(ticker=quote_plus(ticker))}"
    response = requests.get(company_url, headers=auth_headers, verify=False)
    json = response.json().get("results")[0]
    csin = json.get("csin")
    latest_version = (
        json.get("latest_equity_model", {}).get("model_version", {}).get("name", {})
    )

    return (csin, latest_version)


def get_csin_from_ticker(ticker, auth_headers, mds_host=settings.MDS_HOST):
    company_url = f"{mds_host}/{CSIN_URL.format(ticker=quote_plus(ticker))}"
    response = requests.get(company_url, headers=auth_headers)
    json = response.json().get("results")[0]
    csin = json.get("csin")
    return csin


def get_excel_model(ticker, config):
    auth_headers = get_api_headers(config.canalyst_api_key)
    mds_host = config.mds_host
    wp_host = config.wp_host
    default_dir = config.default_dir

    csin = get_csin_from_ticker(ticker, auth_headers, mds_host)

    client = GraphqlClient(endpoint=f"{wp_host}/model-workbooks")

    # Create the query string and variables required for the request.
    query = """
    query driversWorksheetByCSIN($csin: ID!) {
        modelSeries(id: $csin) {
        latestModel {
            id
            name
            publishedAt
            variantsByDimensions(
                driversWorksheets: [STANDARD_FCF],
                periodOrder: [CHRONOLOGICAL],
            ) {
            id
            downloadUrl
            variantDimensions {
                driversWorksheets
                periodOrder
            }
            }
        }
        }
    }
    """
    variables = {"csin": csin}

    # Synchronous request
    data = client.execute(
        query=query, variables=variables, headers=auth_headers, verify=False
    )
    url = data["data"]["modelSeries"]["latestModel"]["variantsByDimensions"][0][
        "downloadUrl"
    ]
    file_ticker = ticker.split(" ")[0]
    file_name = data["data"]["modelSeries"]["latestModel"]["name"]
    file_name = f"{default_dir}/DATA/{file_ticker}/{file_name}.xlsx"

    if path.exists(file_name):
        return file_name

    os.makedirs(f"{default_dir}/DATA/{file_ticker}/", exist_ok=True)
    print("Saved to: " + file_name)
    r = requests.get(url, headers=auth_headers, verify=False)

    with open(file_name, "wb") as f:
        f.write(r.content)

    return


def create_pdf_from_dot(dot_file):
    pdf_file = dot_file.replace(".dot", ".pdf")
    import graphviz

    s = graphviz.Source.from_file(dot_file)
    s.render()
    return pdf_file


def get_forecast_url(csin, latest_version, mds_host=settings.MDS_HOST):
    return (
        f"{mds_host}/api/equity-model-series/{csin}/equity-models/"
        f"{latest_version}/forecast-periods/"
    )


def dot_parse(
    default_df, dot_file="", ticker="", auth_headers="", config=None, s3_client=None
):
    file_ticker = ticker.split(" ")[0]
    if not config:
        config = settings.CONFIG or resolve_config()
    if not s3_client:
        s3_client = Getter(config)

    a_file = s3_client.get_file_from_s3(f"DATA/{file_ticker}/drivers.dot")

    if a_file is None:
        dot_file = create_drivers_dot("net-revenue", auth_headers, ticker)
        a_file = open(dot_file)
    else:
        a_file = str(a_file).split("\\n")

    keys = []
    values = []

    if type(a_file) == list:
        a_file = iter(a_file)

    next(a_file)

    for line in a_file:
        if "}" not in line:
            if "->" in line:
                key = line.split("->")[0]
                key = key.replace("  ", "")
                key = key.split("|")[0]
                key = key.replace('"', "")
                value = line.split("->")[1]
                value = value.split("|")[0]
                value = value.replace("\\n", "")
                value = value.replace('"', "")
                value = value.replace(" ", "")
                value = value.replace("label=", "")
                keys.append(key)
                values.append(value)

    data_tuples = list(zip(keys, values))
    df = pd.DataFrame(data_tuples, columns=["dependent", "precedent"])
    df = df.reset_index()

    df1 = pd.merge(
        df, default_df, how="inner", left_on="precedent", right_on="time_series_slug"
    )[
        ["index", "precedent", "time_series_name", "time_series_description"]
    ].sort_values(
        "index"
    )
    df2 = pd.merge(
        df, default_df, how="inner", left_on="dependent", right_on="time_series_slug"
    )[
        ["index", "dependent", "time_series_name", "time_series_description"]
    ].sort_values(
        "index"
    )
    df = pd.merge(
        df1, df2, how="inner", left_on="index", right_on="index"
    )  # [['precedent','dependent','precedent_time_series_name','precedent_time_series_description','dependent_time_series_name','dependent_time_series_description']]
    cols = []
    for col in df.columns:
        col = col.replace("_x", "_dependent")
        col = col.replace("_y", "_precedent")
        cols.append(col)
    df.columns = cols
    df = df[
        [
            "index",
            "precedent",
            "dependent",
            "time_series_name_precedent",
            "time_series_description_precedent",
            "time_series_name_dependent",
            "time_series_description_dependent",
        ]
    ]
    df = df.groupby(["precedent", "dependent"]).first().reset_index()
    df = df.sort_values("index")

    return df


def CompList(comp_name=""):
    return COMP_LISTS.get(comp_name, "")


def get_forecast_url_data(res_dict, ticker, auth_headers):
    list_out = []
    url = res_dict["self"]

    res_loop = get_request(url, auth_headers)
    url = res_loop.json()["data_points"]

    res_loop = get_request(url, auth_headers)
    # try:
    next_url = res_loop.json()["next"]
    # except:
    #    next_url = None

    while url is not None:

        res_data = res_loop.json()["results"]
        dict_out = {}

        for res_data_dict in res_data:
            dict_out["time_series_slug"] = res_data_dict["time_series"]["slug"]
            dict_out["time_series_name"] = res_data_dict["time_series"]["names"][0]
            dict_out["time_series_description"] = res_data_dict["time_series"][
                "description"
            ]
            dict_out["category_slug"] = res_data_dict["time_series"]["category"]["slug"]
            dict_out["category"] = res_data_dict["time_series"]["category"][
                "description"
            ]  # ?
            dict_out["category_type_slug"] = res_data_dict["time_series"]["category"][
                "type"
            ]["slug"]
            dict_out["category_type_name"] = res_data_dict["time_series"]["category"][
                "type"
            ]["name"]
            dict_out["unit_description"] = res_data_dict["time_series"]["unit"][
                "description"
            ]
            dict_out["unit_type"] = res_data_dict["time_series"]["unit"][
                "unit_type"
            ]  # ?
            dict_out["unit_symbol"] = res_data_dict["time_series"]["unit"]["symbol"]
            dict_out["period_name"] = res_data_dict["period"]["name"]
            dict_out["period_duration_type"] = res_data_dict["period"][
                "period_duration_type"
            ]
            dict_out["period_start_date"] = res_data_dict["period"]["start_date"]
            dict_out["period_end_date"] = res_data_dict["period"]["end_date"]
            dict_out["value"] = res_data_dict["value"]
            dict_out["ticker"] = ticker
            df = pd.DataFrame.from_dict(dict_out, orient="index").T
            list_out.append(df)
        url = next_url
        try:
            res_loop = get_request(url, auth_headers)
            next_url = res_loop.json()["next"]
        except:
            url = None
    return pd.concat(list_out)


def get_scenario_url_data(
    res_dict,
    ticker,
    auth_headers,
    default_dir=settings.DEFAULT_DIR,
):
    log = LogFile(default_dir=default_dir)
    url = res_dict["self"]

    res_loop = get_request(url, auth_headers)
    url = res_loop.json()["data_points"]

    res_loop = get_request(url, auth_headers)

    try:
        res_data = res_loop.json()["results"]
    except:
        log.write("Scenario timeout: " + url)
        return

    url = res_loop.json()["next"]
    list_out = []
    while url is not None:
        dict_out = {}
        for res_data_dict in res_data:
            dict_out["time_series_slug"] = res_data_dict["time_series"]["slug"]
            dict_out["time_series_name"] = res_data_dict["time_series"]["names"][0]
            dict_out["time_series_description"] = res_data_dict["time_series"][
                "description"
            ]
            dict_out["category_slug"] = res_data_dict["time_series"]["category"]["slug"]
            dict_out["category_type_slug"] = res_data_dict["time_series"]["category"][
                "type"
            ]["slug"]
            dict_out["category_type_name"] = res_data_dict["time_series"]["category"][
                "type"
            ]["name"]
            dict_out["unit_description"] = res_data_dict["time_series"]["unit"][
                "description"
            ]
            dict_out["unit_symbol"] = res_data_dict["time_series"]["unit"]["symbol"]
            dict_out["period_name"] = res_data_dict["period"]["name"]
            dict_out["period_duration_type"] = res_data_dict["period"][
                "period_duration_type"
            ]
            dict_out["period_start_date"] = res_data_dict["period"]["start_date"]
            dict_out["period_end_date"] = res_data_dict["period"]["end_date"]
            dict_out["value"] = res_data_dict["value"]
            dict_out["ticker"] = ticker
            df = pd.DataFrame.from_dict(dict_out, orient="index").T
            list_out.append(df)
        try:
            res_loop = get_request(url, auth_headers)
            res_data = res_loop.json()["results"]
            url = res_loop.json()["next"]
        except:
            url = None
    return pd.concat(list_out)


def get_request(url, headers):
    response = requests.get(url=url, headers=headers, verify=False)
    if check_throttle_limit(response):
        response = requests.get(url=url, headers=headers, verify=False)
    return response


def check_throttle_limit(response):
    if response.status_code == 429:
        timer = int(response.headers.get("Retry-After"))
        print(f"Please wait...Retrying request in {timer} seconds")
        time.sleep(timer)
        return True
    return False


def scandir(dir, ext):  # dir: str, ext: list
    subfolders, files = [], []

    for f in os.scandir(dir):
        if f.is_dir():
            subfolders.append(f.path)
        if f.is_file():
            if os.path.splitext(f.name)[1].lower() in ext:
                files.append(f.path)

    # to iterate is human, to recurse is divine
    for dir in list(subfolders):
        sf, f = scandir(dir, ext)
        subfolders.extend(sf)
        files.extend(f)
    return subfolders, files


def select_distinct(df, col):
    d = {}
    d[col] = df[col].unique()
    d = pd.DataFrame(d)
    return d


def create_drivers_dot(time_series, auth_headers, ticker, config=None, s3_client=None):
    if time_series == "":
        time_series = "net-revenue"

    file_ticker = ticker.split(" ")[0]

    path_name = f"DATA/{file_ticker}/"

    if not config:
        config = settings.CONFIG or resolve_config()
    if not s3_client:
        s3_client = Getter(config)

    if time_series == "net-revenue":

        dot_file = s3_client.get_file_from_s3(f"DATA/{file_ticker}/drivers.dot")
        if dot_file is not None:
            if not os.path.exists(f"{config.default_dir}/{path_name}"):
                os.makedirs(f"{config.default_dir}/{path_name}")
            file_name = f"{config.default_dir}/{path_name}" + "drivers.dot"
            open(file_name, "wb").write(dot_file)
            return file_name

    if not os.path.exists(f"{config.default_dir}/{path_name}"):
        os.makedirs(f"{config.default_dir}/{path_name}")

    path_name = f"{config.default_dir}/{path_name}/drivers.dot"

    if os.path.exists(path_name) and time_series == "net-revenue":
        return path_name
    else:
        csin, version = get_company_info_from_ticker(
            ticker, auth_headers, config.mds_host
        )
        url = (
            f"{config.mds_host}/api/equity-model-series/{csin}/equity-models/"
            f"{version}/time-series/{time_series}/forecast-data-points/"
        )

        r = requests.get(url, headers=auth_headers, verify=False)
        json = r.json().get("results")[0]
        name = json["period"]["name"]
        url = (
            f"{config.mds_host}/api/equity-model-series/{csin}/equity-models/"
            f"{version}/time-series/{time_series}/forecast-data-points/{name}/"
            f"drivers/?format=dot"
        )

        print(url)
        r = requests.get(url, headers=auth_headers, verify=False)
        open(path_name, "wb").write(r.content)
        return path_name
    return


def get_model_url(ticker, wp_host=settings.WP_HOST):
    file_ticker = ticker.split(" ")[0]
    url = f"{wp_host}/files/search?query={file_ticker}"
    return url


# retrieve latest equity model from API
def get_model_info(
    ticker,
    auth_headers,
    default_dir=settings.DEFAULT_DIR,
    mds_host=settings.MDS_HOST,
):
    log = LogFile(default_dir=default_dir)
    model_info = {}

    company_url = f"{mds_host}/{CSIN_URL.format(ticker=quote_plus(ticker))}"

    response = requests.get(company_url, headers=auth_headers, verify=False)
    try:
        json = response.json().get("results")[0]
    except:
        log.write(
            f"Candas: Error with {company_url}.  Perhaps you need the "
            "correct Canalyst ticker including exchange."
        )
        return
    csin = json.get("csin")
    company = json.get("company", {}).get("name")
    latest_version = (
        json.get("latest_equity_model", {}).get("model_version").get("name")
    )
    earnings_update_type = json.get("latest_equity_model", {}).get(
        "earnings_update_type"
    )

    periods_url = f"{mds_host}/{PERIODS_URL.format(csin=csin, version=latest_version)}"
    periods_response = requests.get(periods_url, headers=auth_headers, verify=False)
    periods_json = periods_response.json().get("results")
    periods = [
        period.get("name") for period in periods_json[:HISTORICAL_PERIODS_TO_FETCH]
    ]

    model_info[ticker] = (csin, company, latest_version, periods, earnings_update_type)

    return_string = (
        f"{mds_host}/api/equity-model-series/{csin}/equity-models/"
        f"{latest_version}/historical-data-points/?page_size=500"
    )

    return return_string, model_info


def write_json(json_data, file_name):
    with open(file_name, "w") as f:
        json.dump(json_data, f)


def crawl_company_pages(
    next_url,
    ticker,
    key_name,
    next_name,
    auth_headers,
    default_dir=settings.DEFAULT_DIR,
):
    file_ticker = ticker.split(" ")[0]

    page_number = 1
    while next_url is not None:
        response, next_url = get_json(next_url, key_name, next_name, auth_headers)
        files_path = f"{default_dir}/DATA/{file_ticker}"
        os.makedirs(files_path, exist_ok=True)
        file_name = f"{files_path}/{page_number}.json"
        write_json(response, file_name)
        page_number += 1
    return

    # get api json data


def get_json(url, key_name, next_name, auth_headers):
    response = requests.get(url, headers=auth_headers, verify=False)
    res = response.json().get(key_name)
    next_url = response.json().get(next_name)
    return res, next_url


def read_json(ticker, default_dir=settings.DEFAULT_DIR):
    ticker = ticker.split(" ")[0]
    files_path = f"{default_dir}/DATA/{ticker}/"
    arr = os.listdir(files_path)
    dict_js = {}
    for file_name in arr:
        if ".json" in file_name:
            file_name = f"{default_dir}/DATA/{ticker}/{file_name}"
            with open(file_name, "r") as j:
                contents = json.loads(j.read())
            dict_js[file_name] = contents
            os.remove(file_name)
    return dict_js


def json_to_df(dict_json, ticker):
    json_key_list = sorted(dict_json.keys())
    list_out = []
    for key in json_key_list:
        for i in range(len(dict_json[key])):
            dict_out = {}
            content = dict_json[key][i]
            dict_out["ticker"] = ticker
            dict_out["period_name"] = content["period"]["name"]
            dict_out["period_duration_type"] = content["period"]["period_duration_type"]
            dict_out["period_start_date"] = content["period"]["start_date"]
            dict_out["period_end_date"] = content["period"]["end_date"]
            dict_out["category"] = content["time_series"]["category"]["description"]
            dict_out["category_type_slug"] = content["time_series"]["category"]["type"][
                "slug"
            ]  # ?
            dict_out["time_series_slug"] = content["time_series"][
                "slug"
            ]  # api to use slugs vs names? ... can map two or more slugs to the same name ...
            dict_out["time_series_name"] = content["time_series"]["names"][
                0
            ]  # can't apply the same name to more than one time series model - an excel thing
            dict_out["category_type_slug"] = content["time_series"]["category"]["type"][
                "slug"
            ]  # financial or operating stats or other
            dict_out["category_type_name"] = content["time_series"]["category"]["type"][
                "name"
            ]  # financial or operating stats or other
            dict_out["time_series_description"] = content["time_series"][
                "description"
            ]  # for use when we have not applied MO names
            dict_out["unit_description"] = content["time_series"]["unit"]["description"]
            dict_out["unit_symbol"] = content["time_series"]["unit"]["symbol"]
            dict_out["unit_type"] = content["time_series"]["unit"]["unit_type"]
            dict_out["value"] = content["value"]
            df = pd.DataFrame.from_dict(dict_out, orient="index").T
            list_out.append(df)

    df_data = pd.concat(list_out)
    return df_data


def mrq_df(df):
    df = df[df["is_historical"] == True]

    n_size = df.groupby(["ticker", "time_series_name"]).size().mean()
    df["n"] = df.groupby(["ticker", "time_series_name"]).cumcount() - (n_size - 1)
    df["n"] = df["n"].astype(int)

    cols = list(df.groupby("n").first().reset_index()["n"])

    df = pd.pivot_table(
        df,
        values="value",
        index=[
            "name_index",
            # "model_name",
            "ticker",
            "category",
            "time_series_name",
            "time_series_description",
            "is_driver",
            "MRQ",
        ],
        columns=["n"],
        aggfunc=np.sum,
    ).reset_index()
    list_cols = []

    name_cols = [
        "name_index",
        # "model_name",
        "ticker",
        "category",
        "time_series_name",
        "time_series_description",
        "is_driver",
        "MRQ",
    ]

    col_names = []
    for col in df.columns:
        if col not in name_cols:
            col = "MRQ" + str(col)
            col_names.append(col)
    df.columns = name_cols + col_names
    return df.sort_values("name_index")


def pivot_df(df, mrq_notation):

    if mrq_notation == True:
        df = mrq_df(df)
        return df

    df = pd.pivot_table(
        df,
        values="value",
        index=[
            # "model_name",
            "ticker",
            "name_index",
            "category",
            "time_series_name",
            "time_series_description",
            "is_driver",
        ],
        columns=["period_name_sorted"],
        aggfunc=np.sum,
    ).reset_index()

    df = df.sort_values("name_index")
    return df


def refresh_cache(config: Config = None, log: LogFile = None, s3_client: Getter = None):
    if not config:
        config = settings.CONFIG or resolve_config()
    if not log:
        log = LogFile(default_dir=config.default_dir)
    if not s3_client:
        s3_client = Getter(config=config)

    folder = "DATA"
    subfolders, files = scandir(folder, [".csv"])
    for filename in files:
        try:
            print("Candas: refresh " + filename)
            os.remove(filename)
            s3_client.get_csv_from_s3(filename)
        except:
            log.write("Candas: get error for " + filename)
    return


def get_cache_tickers(default_dir=settings.DEFAULT_DIR):

    folder = "DATA"
    folder = f"{default_dir}/{folder}"
    subfolders, files = scandir(folder, [".xlsx"])
    ticker_list = []
    for filename in files:
        if "data" not in filename:
            ticker_list.append(filename)
    return ticker_list


def get_candas_ticker_list(ticker="", config=None):
    list_files = []
    ticker = ticker.split(" ")[0]

    if not config:
        config = settings.CONFIG or resolve_config()

    session = Session(
        aws_access_key_id=config.s3_access_key_id,
        aws_secret_access_key=config.s3_secret_key,
    )
    s3 = session.resource("s3")
    your_bucket = s3.Bucket("candas")
    for s3_file in your_bucket.objects.all():
        str_key = s3_file.key
        try:
            str_key = str_key.split("/")[1]
            if str_key != "DATA":
                list_files.append(str_key)
        except:
            continue
    list_files = list(set(list_files))
    if ticker != "":
        return list_files.count(ticker)
    else:
        return list_files


def find_in_list(feature_list, search_term):
    return_list = []
    for item in feature_list:
        if search_term.lower() in item.lower():
            return_list.append(item)
    return return_list


# helper function ... simply df loc
def df_filter(df, col, feature_list):
    if col == "":
        col = "time_series_name"
    if type(feature_list) == str:
        df = df.loc[df[col].str.contains(feature_list)]
        return df
    df = df.loc[df[col].isin(feature_list)]
    return df


def plot_all(
    df,
    index_col,
    group_col,
    value_col,
    title_text,
    plot_kind="line",
    allow_na=False,
    n="",
):

    df[value_col] = df[value_col].astype(float)
    df = df.pivot(index=index_col, columns=group_col, values=value_col)
    if allow_na == False:
        df = df.dropna()
    if n != "":
        df.head(n)
    plt = df.plot(title=title_text, kind=plot_kind)
    plt.legend(loc="center left", bbox_to_anchor=(1.0, 0.5))
    plt.plot()


def plot_all_labels(
    df, index_col, group_col, value_col, title_text, plot_kind="line", labels=None
):
    df[value_col] = df[value_col].astype(float)
    df = df.pivot(index=index_col, columns=group_col, values=value_col)
    # df = df.dropna()
    return df.plot(title=title_text, kind=plot_kind, labels=labels)


def calendar_quarter(df, col, datetime=True):
    pd.set_option("mode.chained_assignment", None)
    # translate a date into sort-able and group-able YYYY-mm format.
    df[col] = pd.to_datetime(df[col])

    df[col + "shift"] = df[col] + pd.Timedelta(days=-12)

    df[col + "_CALENDAR_QUARTER"] = df[col + "shift"].dt.to_period("Q")

    df = df.drop(columns=[col + "shift"])
    df[col + "_CALENDAR_QUARTER"] = df[col + "_CALENDAR_QUARTER"].astype(str)

    return df


COMP_LISTS = {}
COMP_LISTS["Apparel"] = [
    "FL US",
    "BGFV US",
    "BKE US",
    "BURL US",
    "MIK US",
    "ULTA US",
    "TSCO US",
    "BOOT US",
    "DXLG US",
    "SCVL US",
]
COMP_LISTS["SAAS"] = [
    "ASAN US",
    "AYX US",
    "AMSWA US",
    "PLAN US",
    "APPF US",
    "TEAM US",
    "ADSK US",
    "AVLR US",
    "BILL US",
    "BLKB US",
    "BL US",
    "BOX US",
    "NET US",
    "COUP US",
    "CRWD US",
    "DDOG US",
    "DOCU US",
    "DOMO US",
    "DBX US",
    "DT US",
    "ESTC US",
    "HUBS US",
    "JAMF US",
    "FROG US",
    "MDLA US",
    "NCNO US",
    "NEWR US",
    "OKTA US",
    "PD US",
    "PAYC US",
    "PCTY US",
    "PING US",
    "PS US",
    "QTWO US",
    "RNG US",
    "CRM US",
    "NOW US",
    "SHOP CN",
    "WORK US",
    "SMAR US",
    "SNOW US",
    "SUMO US",
    "TLND US",
    "TENB US",
    "TWLO US",
    "VEEV US",
    "WIX US",
    "WDAY US",
    "ZEN US",
    "ZM US",
    "ZI US",
    "ZS US",
]
